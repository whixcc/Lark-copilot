import os
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
import time
import json
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import glob
import re
import sys
import logging
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.common.exceptions import SessionNotCreatedException
from datetime import datetime, timedelta

# 配置文件路径
CONFIG_FILE = 'config2.json'

# 定义webdriver路径 (保持不变，headless模式不需要特定路径)
def get_driver_path(browser):
    if getattr(sys, 'frozen', False):
        base_path = os.path.join(sys._MEIPASS, 'drivers')
    else:
        base_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'drivers')
    if browser == 'chrome':
        return os.path.join(base_path, 'chromedriver.exe')
    elif browser == 'firefox':
        return os.path.join(base_path, 'geckodriver.exe')
    elif browser == 'edge':
        return os.path.join(base_path, 'msedgedriver.exe')
    return None

# 配置日志
logging.basicConfig(filename='隐患提醒.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class ConfigWindow:
    def __init__(self, parent=None):
        self.window = tk.Toplevel(parent) if parent else tk.Tk()
        self.window.title("配置信息")
        self.window.geometry("550x300")
        self.window.resizable(False, False)

        if parent:
            parent.iconify()
            self.window.protocol("WM_DELETE_WINDOW", lambda: self.on_closing(parent))
        else:
            self.window.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.config = self.load_config()

        self.downloads_folder_var = tk.StringVar()
        self.roster_path_var = tk.StringVar()

        if self.config:
            self.downloads_folder_var.set(self.config.get('downloads_folder', ''))
            self.roster_path_var.set(self.config.get('roster_path', ''))

        self.create_widgets()

    def create_widgets(self):
        tk.Label(self.window, text="PIN码:", font=('Arial', 10)).grid(row=0, column=0, padx=10, pady=5, sticky='w')
        self.pin_entry = tk.Entry(self.window, width=50)
        self.pin_entry.grid(row=0, column=1, padx=10, pady=5, sticky='ew')

        tk.Label(self.window, text="Webhook URL:", font=('Arial', 10)).grid(row=1, column=0, padx=10, pady=5, sticky='w')
        self.webhook_entry = tk.Entry(self.window, width=70)
        self.webhook_entry.grid(row=1, column=1, padx=10, pady=5, sticky='ew')

        tk.Label(self.window, text="下载文件夹路径:", font=('Arial', 10)).grid(row=2, column=0, padx=10, pady=5, sticky='w')
        self.downloads_folder_label = tk.Label(self.window, textvariable=self.downloads_folder_var, width=60, anchor='w', relief='groove', borderwidth=1)
        self.downloads_folder_label.grid(row=2, column=1, padx=10, pady=5, sticky='ew')
        ttk.Button(self.window, text="选择", command=self.browse_downloads_folder).grid(row=2, column=2, padx=5, pady=5)

        tk.Label(self.window, text="花名册路径:", font=('Arial', 10)).grid(row=3, column=0, padx=10, pady=5, sticky='w')
        self.roster_path_label = tk.Label(self.window, textvariable=self.roster_path_var, width=60, anchor='w', relief='groove', borderwidth=1)
        self.roster_path_label.grid(row=3, column=1, padx=10, pady=5, sticky='ew')
        ttk.Button(self.window, text="选择", command=self.browse_roster_path).grid(row=3, column=2, padx=5, pady=5)

        # 添加浏览器选择
        tk.Label(self.window, text="选择浏览器:", font=('Arial', 10)).grid(row=4, column=0, padx=10, pady=5, sticky='w')
        self.browser_var = tk.StringVar()
        browser_combobox = ttk.Combobox(
            self.window,
            textvariable=self.browser_var,
            values=["Chrome", "Firefox", "Edge"],
            state="readonly"
        )
        browser_combobox.grid(row=4, column=1, padx=10, pady=5, sticky='ew')
        browser_combobox.set(self.config.get('browser', 'chrome') if self.config else 'chrome')

        if self.config:
            self.pin_entry.insert(0, self.config.get('pin_code', ''))
            self.webhook_entry.insert(0, self.config.get('webhook_url', ''))

        ttk.Button(
            self.window,
            text="保存配置",
            command=self.save
        ).grid(row=5, column=0, columnspan=3, pady=15)

        self.status_label = tk.Label(self.window, text="", fg="green")
        self.status_label.grid(row=6, column=0, columnspan=3, pady=5)

        self.window.columnconfigure(1, weight=1)

    def browse_downloads_folder(self):
        folder_selected = filedialog.askdirectory(title="选择下载文件夹")
        if folder_selected:
            self.downloads_folder_var.set(folder_selected)
            logging.info(f"用户选择了下载文件夹: {folder_selected}")

    def browse_roster_path(self):
        file_selected = filedialog.askopenfilename(title="选择花名册文件", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_selected:
            self.roster_path_var.set(file_selected)
            logging.info(f"用户选择了花名册文件: {file_selected}")

    def load_config(self):
        logging.info("加载配置文件...")
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    logging.info("配置文件加载成功。")
                    return config
            except json.JSONDecodeError as e:
                logging.error(f"解析配置文件失败: {e}")
                messagebox.showerror("错误", f"解析配置文件失败: {e}")
                return None
        else:
            logging.info("配置文件不存在。")
            return None

    def save(self):
        pin_code = self.pin_entry.get().strip()
        webhook_url = self.webhook_entry.get().strip()
        downloads_folder = self.downloads_folder_var.get().strip()
        roster_path = self.roster_path_var.get().strip()
        browser = self.browser_var.get().strip().lower()

        if not pin_code or not webhook_url or not downloads_folder or not roster_path or not browser:
            messagebox.showerror("错误", "请填写所有字段！")
            logging.warning("保存配置失败：存在未填写的字段。")
            return

        config = {
            'pin_code': pin_code,
            'webhook_url': webhook_url,
            'downloads_folder': downloads_folder,
            'roster_path': roster_path,
            'browser': browser
        }

        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            self.status_label.config(text="配置已保存！", fg="green")
            logging.info("配置已保存。")
            self.window.after(2000, lambda: self.status_label.config(text=""))
        except Exception as e:
            messagebox.showerror("错误", f"保存配置失败: {e}")
            logging.error(f"保存配置失败: {e}")

    def on_closing(self, parent=None):
        if not os.path.exists(CONFIG_FILE):
            if messagebox.askokcancel("退出", "配置尚未保存，确定要退出吗？"):
                self.window.destroy()
                if parent:
                    parent.deiconify()
                logging.info("配置窗口关闭，未保存配置。")
        else:
            self.window.destroy()
            if parent:
                parent.deiconify()
            logging.info("配置窗口关闭。")

class MainWindow:
    def __init__(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        drivers_dir = os.path.join(script_dir, 'drivers')
        if not os.path.exists(drivers_dir):
            os.makedirs(drivers_dir)

        self.root = tk.Tk()
        self.root.title("隐患和随手拍到期提醒")
        self.root.geometry("350x200")
        self.selected_browser = tk.StringVar(self.root)
        self.selected_browser.set("chrome")

        self.config = self.load_config()
        if len(sys.argv) > 1 and sys.argv[1] == '--auto':
            logging.info("以 --auto 模式启动。")
            self.create_widgets()  # 先创建小部件
            if self.config and 'browser' in self.config:
                self.selected_browser.set(self.config['browser'])
                logging.info(f"自动模式下选择的浏览器: {self.config['browser']}")
            else:
                self.selected_browser.set("chrome")  # 默认浏览器
                logging.info("配置文件中未找到浏览器设置，默认使用 Chrome。")
            self.run_automation()
        else:
            logging.info("以正常模式启动。")
            # 如果不是自动模式，尝试从配置中设置默认浏览器
            if self.config and 'browser' in self.config:
                self.selected_browser.set(self.config['browser'])
            self.create_widgets()
            self.bind_shortcuts()
            self.root.protocol("WM_DELETE_WINDOW", self.close_application)

    def create_widgets(self):
        ttk.Button(
            self.root,
            text="设置",
            command=self.open_config
        ).pack(pady=20, padx=20, fill='x')

        ttk.Button(
            self.root,
            text="运行",
            command=self.run_automation
        ).pack(pady=20, padx=20, fill='x')

        self.status_label = tk.Label(self.root, text="", fg="black")
        self.status_label.pack(pady=10)

    def open_config(self):
        logging.info("打开配置窗口。")
        ConfigWindow(self.root)

    def run_automation(self):
        logging.info("开始执行自动化任务...")
        self.root.iconify()
        self.status_label.config(text="正在运行...", fg="blue")
        self.root.update()

        # 仅在非自动模式下保存浏览器选择
        if not (len(sys.argv) > 1 and sys.argv[1] == '--auto'):
            self.save_browser_selection()

        try:
            self.perform_automation()
            self.status_label.config(text="运行完成！", fg="green")
            logging.info("自动化任务执行完成。")
        except Exception as e:
            error_message = f"运行出错: {str(e)}"
            self.status_label.config(text=error_message, fg="red")
            logging.error(error_message)
            messagebox.showerror("错误", error_message)
        finally:
            self.root.deiconify()

    def bind_shortcuts(self):
        self.root.bind('<Control-Alt-R>', lambda event: self.run_automation())
        self.root.bind('<Control-Alt-S>', lambda event: self.open_config())

    def load_config(self):
        logging.info("加载配置文件...")
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    logging.info("配置文件加载成功。")
                    return config
            except json.JSONDecodeError as e:
                logging.error(f"解析配置文件失败: {e}")
                return None
        else:
            logging.info("配置文件不存在。")
            return None

    def save_browser_selection(self):
        if not self.config:
            self.config = {}
        self.config['browser'] = self.selected_browser.get().lower()
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
            logging.info("浏览器选择已保存到配置文件。")
        except Exception as e:
            logging.error(f"保存浏览器选择失败: {e}")

    def setup_driver(self, browser_name, headless=False):
        logging.info(f"正在设置浏览器驱动: {browser_name}, headless={headless}")
        driver_path = get_driver_path(browser_name)

        try:
            if browser_name == "chrome":
                options = ChromeOptions()
                if headless:
                    options.add_argument('--headless=new')  # 或使用 '--headless'
                    options.add_argument('--disable-gpu')
                    options.add_argument('--no-sandbox')
                    options.add_argument('--window-size=1920,1080')
                if driver_path and os.path.exists(driver_path):
                    service = ChromeService(driver_path)
                    driver = webdriver.Chrome(service=service, options=options)
                    logging.info("使用本地打包的 ChromeDriver。")
                else:
                    driver_path = ChromeDriverManager().install()
                    service = ChromeService(driver_path)
                    driver = webdriver.Chrome(service=service, options=options)
                    logging.info("使用 webdriver-manager 下载的 ChromeDriver。")
            elif browser_name == "firefox":
                options = FirefoxOptions()
                if headless:
                    options.headless = True
                    options.add_argument("--width=1920")
                    options.add_argument("--height=1080")
                if driver_path and os.path.exists(driver_path):
                    service = FirefoxService(driver_path)
                    driver = webdriver.Firefox(service=service, options=options)
                    logging.info("使用本地打包的 GeckoDriver。")
                else:
                    driver_path = GeckoDriverManager().install()
                    service = FirefoxService(driver_path)
                    driver = webdriver.Firefox(service=service, options=options)
                    logging.info("使用 webdriver-manager 下载的 GeckoDriver。")
            elif browser_name == "edge":
                options = EdgeOptions()
                if headless:
                    options.add_argument('--headless=new')  # 或使用 '--headless'
                    options.add_argument('--disable-gpu')
                    options.add_argument('--no-sandbox')
                    options.add_argument('--window-size=1920,1080')
                if driver_path and os.path.exists(driver_path):
                    service = EdgeService(driver_path)
                    driver = webdriver.Edge(service=service, options=options)
                    logging.info("使用本地打包的 EdgeDriver。")
                else:
                    driver_path = EdgeChromiumDriverManager().install()
                    service = EdgeService(driver_path)
                    driver = webdriver.Edge(service=service, options=options)
                    logging.info("使用 webdriver-manager 下载的 EdgeDriver。")
            else:
                raise ValueError(f"不支持的浏览器: {browser_name}")
            return driver
        except SessionNotCreatedException as e:
            logging.warning(f"使用本地驱动创建会话失败: {e}")
            logging.info(f"尝试使用 webdriver-manager 下载 {browser_name} 的驱动。")
            if browser_name == "chrome":
                driver_path = ChromeDriverManager().install()
                service = ChromeService(driver_path)
                driver = webdriver.Chrome(service=service)
            elif browser_name == "firefox":
                driver_path = GeckoDriverManager().install()
                service = FirefoxService(driver_path)
                driver = webdriver.Firefox(service=service)
            elif browser_name == "edge":
                driver_path = EdgeChromiumDriverManager().install()
                service = EdgeService(driver_path)
                driver = webdriver.Edge(service=service)
            logging.info(f"成功使用 webdriver-manager 下载的驱动创建 {browser_name} 会话。")
            return driver
        except Exception as e:
            logging.error(f"初始化 {browser_name} 浏览器驱动失败: {e}")
            raise

    def perform_automation(self):
        logging.info("执行自动化步骤...")
        if not os.path.exists(CONFIG_FILE):
            error_message = "配置文件不存在，请先进行配置"
            logging.error(error_message)
            raise Exception(error_message)

        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
        except json.JSONDecodeError as e:
            error_message = f"解析配置文件失败: {e}"
            logging.error(error_message)
            raise Exception(error_message)

        pin_code = config.get('pin_code')
        webhook_url = config.get('webhook_url')
        downloads_folder = config.get('downloads_folder')
        roster_path = config.get('roster_path')

        if not pin_code or not webhook_url or not downloads_folder or not roster_path:
            error_message = "配置信息不完整"
            logging.error(error_message)
            raise Exception(error_message)

        browser_name = self.selected_browser.get().lower()
        logging.info(f"选择的浏览器: {browser_name}")

        # 根据是否以 --auto 模式启动，决定是否使用 headless
        headless = len(sys.argv) > 1 and sys.argv[1] == '--auto'
        driver = self.setup_driver(browser_name, headless=headless)
        if not driver:
            logging.error(f"未能成功初始化 {browser_name} 浏览器驱动。")
            return

        try:
            driver.get("https://ehs.crland.com.cn/home")
            logging.info(f"在 {browser_name} 中打开网页: https://ehs.crland.com.cn/home")
            wait = WebDriverWait(driver, 20)

            pin_input = wait.until(EC.presence_of_element_located(
                (By.XPATH, '//input[@type="password" and @placeholder="pin码" and contains(@class, "el-input__inner")]')))
            pin_input.send_keys(pin_code)
            time.sleep(1)
            pin_input.send_keys(Keys.RETURN)
            logging.info("PIN码已输入并提交。")

            if not headless:
                driver.maximize_window()
            time.sleep(5)

            # 隐患排查清单导出
            security_check_menu = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "安全检查")))
            security_check_menu.click()
            logging.info("点击安全检查菜单。")
            time.sleep(1)

            special_check_menu = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "专项检查")))
            special_check_menu.click()
            logging.info("点击专项检查菜单。")
            time.sleep(1)

            hidden_danger_list_menu = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "隐患查询列表")))
            hidden_danger_list_menu.click()
            logging.info("点击隐患查询列表菜单。")
            time.sleep(2)

            # 日期选择
            deadline_date_input = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[1]/div/form/div[1]/div[3]/div/div/div/div[1]/div/input')))
            deadline_date_input.click()
            logging.info("点击隐患列表日期选择输入框。")
            time.sleep(0.5)

            today = datetime.now()
            seven_days_ago = today - timedelta(days=7)
            date_range = f"{seven_days_ago.strftime('%Y-%m-%d')} - {today.strftime('%Y-%m-%d')}"
            deadline_date_input.send_keys(Keys.CONTROL + "a")  # 选择所有文本
            deadline_date_input.send_keys(Keys.DELETE)  # 删除选择的文本
            deadline_date_input.send_keys(date_range)
            logging.info(f"已输入隐患列表的日期范围: {date_range}")
            time.sleep(0.5)
            deadline_date_input.send_keys(Keys.ENTER)
            logging.info("隐患列表日期范围输入完成并提交。")
            time.sleep(1)

            driver.find_element(By.TAG_NAME, 'body').click()
            time.sleep(1)

            # 展开查询选项
            expand_options = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[2]/div/div[1]')))
            expand_options.click()
            logging.info("展开隐患列表的查询选项。")
            time.sleep(1)

            danger_status_dropdown = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[1]/div/form/div[3]/div[2]/div[2]/div/div/div')))
            danger_status_dropdown.click()
            logging.info("打开隐患状态下拉菜单。")
            time.sleep(1)

            pending_correction = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[1]/div/form/div[3]/div[2]/div[2]/div/div/div/div[2]/ul[2]/li[1]')))
            pending_correction.click()

            corrected = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[1]/div/form/div[3]/div[2]/div[2]/div/div/div/div[2]/ul[2]/li[2]')))
            corrected.click()
            logging.info("已选择隐患状态：待整改和已整改。")

            driver.find_element(By.TAG_NAME, 'body').click()

            query_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[2]/div/div[2]/button[2]/span')))
            query_button.click()
            logging.info("点击隐患列表查询按钮。")

            export_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[3]/div/div[2]/div/button')))
            export_button.click()
            logging.info("点击隐患列表导出按钮。")

            no_option = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[6]/div[2]/div/div/div[2]/div/p[1]/div/label[2]')))
            no_option.click()
            logging.info("已选择隐患列表导出'否'。")

            final_export_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[6]/div[2]/div/div/div[2]/div/div/button')))
            final_export_button.click()
            logging.info("最终点击隐患列表导出按钮。")
            time.sleep(2)

            close_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[6]/div[2]/div/div/a/i')))
            close_button.click()
            logging.info("点击隐患列表导出窗口的关闭按钮。")

            # 等待模态框消失
            try:
                wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "ivu-modal-wrap")))
            except Exception as e:
                logging.error(f"等待隐患导出模态框消失时出错: {e}")

            # 随手拍清单导出
            safety_snapshot_menu = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "安全随手拍")))
            safety_snapshot_menu.click()
            logging.info("点击安全随手拍菜单。")
            time.sleep(1)

            snapshot_list_menu = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "随手拍查询列表")))
            snapshot_list_menu.click()
            logging.info("点击随手拍查询列表菜单。")
            time.sleep(2)

            # 日期选择
            deadline_date_input = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[1]/div/form/div[1]/div[3]/div/div/div/div[1]/div/input')))
            deadline_date_input.click()
            logging.info("点击随手拍列表日期选择输入框。")
            time.sleep(0.5)

            today = datetime.now()
            seven_days_ago = today - timedelta(days=7)
            date_range = f"{seven_days_ago.strftime('%Y-%m-%d')} - {today.strftime('%Y-%m-%d')}"
            deadline_date_input.send_keys(Keys.CONTROL + "a")  # 选择所有文本
            deadline_date_input.send_keys(Keys.DELETE)  # 删除选择的文本
            deadline_date_input.send_keys(date_range)
            logging.info(f"已输入随手拍列表的日期范围: {date_range}")
            time.sleep(0.5)
            deadline_date_input.send_keys(Keys.ENTER)
            logging.info("随手拍列表日期范围输入完成并提交。")
            time.sleep(1)

            driver.find_element(By.TAG_NAME, 'body').click()
            time.sleep(1)

            danger_status_dropdown = wait.until(EC.element_to_be_clickable(
                (By.XPATH,
                 '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[1]/div/form/div[1]/div[2]/div/div/div')))
            danger_status_dropdown.click()
            logging.info("打开随手拍隐患状态下拉菜单。")
            time.sleep(1)
            pending_correction = wait.until(EC.element_to_be_clickable(
                (By.XPATH,
                 '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[1]/div[1]/div/form/div[1]/div[2]/div/div/div/div[2]/ul[2]/li[2]')))
            pending_correction.click()
            logging.info("随手拍隐患状态已选择待整改。")

            query_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[3]/div/div[2]/button[2]')))
            query_button.click()
            logging.info("点击随手拍查询按钮。")
            time.sleep(1)

            export_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div[5]/div/div[2]/div[1]/div[1]/div[3]/div/div[2]/button[2]')))
            driver.execute_script("arguments[0].click();", export_button)
            logging.info("点击随手拍导出按钮。")
            time.sleep(2)

            wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "ivu-modal-content")))

            no_option = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'ivu-modal-content')]//label[contains(text(), '否')]")))
            driver.execute_script("arguments[0].click();", no_option)
            logging.info("已选择随手拍'否'。")

            time.sleep(1)
            final_export_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'ivu-modal-content')]//button[contains(@class, 'ivu-btn')]")))
            driver.execute_script("arguments[0].click();", final_export_button)
            logging.info("点击随手拍最终导出按钮。")

            time.sleep(2)

            close_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'ivu-modal-content')]//a[contains(@class, 'ivu-modal-close')]")))
            driver.execute_script("arguments[0].click();", close_button)
            logging.info("点击随手拍导出窗口的关闭按钮。")
            time.sleep(10)

        except Exception as e:
            logging.error(f"自动化过程中出错：{e}")
            raise
        finally:
            if 'driver' in locals():
                driver.quit()
            self.process_data_and_send_message(webhook_url, downloads_folder, roster_path)
            self.delete_downloaded_files(downloads_folder)

    def process_data_and_send_message(self, webhook_url, downloads_folder, roster_path):
        logging.info("开始处理数据并发送消息...")

        def get_latest_file(prefix):
            files = [f for f in os.listdir(downloads_folder) if f.startswith(prefix)]
            if not files:
                logging.info(f"未找到以 '{prefix}' 开头的文件。")
                return None
            latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(downloads_folder, x)))
            logging.info(f"找到最新的文件: {latest_file}")
            return os.path.join(downloads_folder, latest_file)

        def read_excel_file(file_path):
            try:
                logging.info(f"读取 Excel 文件: {file_path}")
                if file_path.endswith('.xlsx'):
                    return pd.read_excel(file_path, engine='openpyxl')
                elif file_path.endswith('.xls'):
                    return pd.read_excel(file_path, engine='xlrd')
                else:
                    error_message = f"不支持的文件格式: {file_path}"
                    logging.error(error_message)
                    raise ValueError(error_message)
            except Exception as e:
                logging.error(f"读取文件 {file_path} 时出错: {e}")
                return None

        roster_df = read_excel_file(roster_path)
        if roster_df is None:
            logging.warning("花名册读取失败，无法发送消息。")
            return
        name_to_id = {str(row['姓名']).strip(): str(row['员工ID']).strip() for index, row in roster_df.iterrows()}
        logging.info("已加载花名册，创建姓名到员工ID的映射。")

        def send_webhook_message(combined_df, webhook_url):
            if combined_df.empty:
                message = "今日无到期隐患"
                logging.info("今日无到期隐患，无需发送消息。")
            else:
                message_lines = ["以下是今日即将到期隐患，下班前务必整改完成并督促复验："]
                for index, row in combined_df.iterrows():
                    responsible_person = str(row['整改责任人']).strip()
                    employee_id = name_to_id.get(responsible_person, responsible_person)
                    if employee_id != responsible_person:
                        message_lines.append(f"{index + 1}、{row['隐患描述']}：<at user_id=\"{employee_id}\"></at>")
                    else:
                        message_lines.append(f"{index + 1}、{row['隐患描述']}：{responsible_person}")
                message = "\n".join(message_lines)

            payload = {
                "msg_type": "text",
                "content": {"text": message}
            }

            try:
                logging.info(f"发送消息到 Webhook: {webhook_url}")
                response = requests.post(webhook_url, json=payload)
                response.raise_for_status()
                logging.info("消息发送成功！")
                logging.info(f"发送的内容：\n{message}")
            except requests.exceptions.RequestException as e:
                logging.error(f"发送消息时出错：{e}")

        hidden_danger_path = get_latest_file("隐患列表")
        hidden_danger_df = pd.DataFrame()
        if hidden_danger_path and os.path.exists(hidden_danger_path):
            hidden_danger_df = read_excel_file(hidden_danger_path)

        snapshot_path = get_latest_file("随手拍")
        snapshot_df = pd.DataFrame()
        if snapshot_path and os.path.exists(snapshot_path):
            snapshot_df = read_excel_file(snapshot_path)

        if not snapshot_df.empty:
            rename_columns = {
                '隐患描述及编号': '隐患描述',
                '整改截至日期': '整改截止日期',
                '隐患整改人': '整改责任人',
            }
            snapshot_df.rename(columns=rename_columns, inplace=True)
            snapshot_df = snapshot_df[list(rename_columns.values())]

        combined_df = pd.concat([hidden_danger_df, snapshot_df], ignore_index=True)

        final_columns = ['隐患描述', '隐患位置', '整改截止日期', '整改责任人', '复验负责人']
        for col in final_columns:
            if col not in combined_df.columns:
                combined_df[col] = None

        combined_df = combined_df[final_columns]

        combined_file_path = os.path.join(downloads_folder, "合并后的隐患列表.xlsx")
        try:
            combined_df.to_excel(combined_file_path, index=False, engine='openpyxl')
            logging.info(f"合并后的数据保存到: {combined_file_path}")
        except Exception as e:
            logging.error(f"保存合并后的数据出错: {e}")

        try:
            wb = load_workbook(combined_file_path)
            ws = wb.active

            column_widths = {'A': 38, 'B': 35, 'C': 15, 'D': 15, 'E': 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row[0].row].height = 18

            for cell in ws["1:1"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            center_columns = ['C', 'D', 'E']
            for col in center_columns:
                for cell in ws[col]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(combined_file_path)
            logging.info(f"格式化后的文件已保存到: {combined_file_path}")
        except Exception as e:
            logging.error(f"格式化并保存文件出错: {e}")

        send_webhook_message(combined_df, webhook_url)

    def delete_downloaded_files(self, directory):
        logging.info(f"开始清理下载文件夹: {directory}")
        patterns = [
            re.compile(r'^合并后的隐患列表[\s]*[\(（]?\d*[\)）]?\.xls[x]?$'),
            re.compile(r'^隐患列表[\s]*[\(（]?\d*[\)）]?\.xls[x]?$'),
            re.compile(r'^随手拍[\s]*[\(（]?\d*[\)）]?\.xls[x]?$'),
        ]

        if not os.path.exists(directory):
            logging.warning(f"路径不存在: {directory}")
            return

        for file_name in os.listdir(directory):
            file_path = os.path.join(directory, file_name)
            if os.path.isfile(file_path):
                for pattern in patterns:
                    if pattern.match(file_name):
                        try:
                            os.remove(file_path)
                            logging.info(f"已删除文件: {file_path}")
                        except Exception as e:
                            logging.error(f"无法删除文件: {file_path}, 错误: {e}")
                        break

    def close_application(self):
        logging.info("关闭应用程序...")
        self.root.destroy()

    def run(self):
        if not os.path.exists(CONFIG_FILE):
            ConfigWindow(self.root)
        self.root.mainloop()

if __name__ == "__main__":
    app = MainWindow()
    app.run()
