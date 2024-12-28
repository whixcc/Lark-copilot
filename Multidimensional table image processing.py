from PIL import JpegImagePlugin
JpegImagePlugin._getmp = lambda x: None  # 忽略MPO多图片对象

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog, messagebox

# 默认配置
config = {
    "image_width": 70,
    "image_height": 30,
    "column_width": 10,
    "row_height": 40
}

def find_image_path(folder_path, image_name):
    """在文件夹及其子文件夹中查找图片路径"""
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file == image_name:
                full_path = os.path.join(root, file)
                if is_valid_image(full_path):  # 确保找到的文件是有效图片
                    return full_path
    return None

def is_valid_image(image_path):
    """检查文件是否为有效图片类型"""
    valid_extensions = ['.jpg', '.jpeg', '.png']
    file_extension = os.path.splitext(image_path)[1].lower()
    return file_extension in valid_extensions

def adjust_cell_size(ws, col, row):
    """根据用户配置调整单元格尺寸"""
    ws.column_dimensions[col].width = config["column_width"]
    ws.row_dimensions[row].height = config["row_height"]

def insert_image_to_cell(ws, row, col, image_path):
    """在指定单元格中插入图片"""
    try:
        img = Image(image_path)
        # 根据用户设置调整图片大小
        img.width, img.height = config["image_width"], config["image_height"]
        col_letter = ws.cell(row=row, column=col).column_letter
        adjust_cell_size(ws, col_letter, row)  # 调整单元格大小
        ws.add_image(img, ws.cell(row=row, column=col).coordinate)  # 插入图片
    except Exception as e:
        print(f"无法插入图片 {image_path}: {e}")

def replace_and_insert_images(excel_path, folder_path):
    """替换Excel中的图片文件名为路径，并插入对应的图片"""
    output_path = os.path.join(os.path.dirname(excel_path), "Updated_" + os.path.basename(excel_path))
    print(f"保存的Excel文件路径: {output_path}")
    try:
        df = pd.read_excel(excel_path)
        wb = load_workbook(excel_path)
        ws = wb.active

        for col_index, col in enumerate(df.columns, start=1):
            for row_index, cell_value in enumerate(df[col], start=2):
                if not isinstance(cell_value, str) or not any(ext in cell_value for ext in ['.jpg', '.jpeg', '.png']):
                    continue
                image_name = cell_value.split(',')[0].strip()
                image_path = find_image_path(folder_path, image_name)
                if image_path and is_valid_image(image_path):
                    # 清除单元格内容
                    ws.cell(row=row_index, column=col_index).value = None
                    # 插入图片
                    insert_image_to_cell(ws, row_index, col_index, image_path)

        wb.save(output_path)
        print(f"Excel文件已保存: {output_path}")
        messagebox.showinfo("完成", f"处理完成！文件已保存到：\n{output_path}")
    except Exception as e:
        messagebox.showerror("处理错误", f"处理时发生错误：\n{e}")

def select_excel():
    """选择Excel文件"""
    global excel_path
    excel_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel文件", "*.xlsx")])
    excel_label.config(text=f"已选择: {excel_path}")

def select_folder():
    """选择图片文件夹"""
    global folder_path
    folder_path = filedialog.askdirectory(title="选择包含图片的文件夹")
    folder_label.config(text=f"已选择: {folder_path}")

def update_config():
    """更新配置"""
    try:
        config["image_width"] = int(image_width_entry.get())
        config["image_height"] = int(image_height_entry.get())
        config["column_width"] = int(column_width_entry.get())
        config["row_height"] = int(row_height_entry.get())
        messagebox.showinfo("设置完成", "配置已更新！")
    except ValueError:
        messagebox.showerror("输入错误", "请输入有效的数字！")

def start_processing():
    """开始处理"""
    if not excel_path or not folder_path:
        messagebox.showwarning("警告", "请先选择Excel文件和图片文件夹")
        return
    replace_and_insert_images(excel_path, folder_path)

# 创建主窗口
root = tk.Tk()
root.title("品质整改表整理")
root.geometry("500x500")

# Excel选择
excel_label = tk.Label(root, text="请选择Excel文件")
excel_label.pack(pady=10)
excel_button = tk.Button(root, text="选择Excel文件", command=select_excel)
excel_button.pack()

# 图片文件夹选择
folder_label = tk.Label(root, text="请选择图片文件夹")
folder_label.pack(pady=10)
folder_button = tk.Button(root, text="选择图片文件夹", command=select_folder)
folder_button.pack()

# 设置区域
settings_label = tk.Label(root, text="设置图片和表格尺寸", font=("Arial", 12))
settings_label.pack(pady=10)

frame = tk.Frame(root)
frame.pack(pady=5)

tk.Label(frame, text="图片宽度(px):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
image_width_entry = tk.Entry(frame, width=10)
image_width_entry.grid(row=0, column=1, padx=5, pady=5)
image_width_entry.insert(0, str(config["image_width"]))

tk.Label(frame, text="图片高度(px):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
image_height_entry = tk.Entry(frame, width=10)
image_height_entry.grid(row=1, column=1, padx=5, pady=5)
image_height_entry.insert(0, str(config["image_height"]))

tk.Label(frame, text="列宽:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
column_width_entry = tk.Entry(frame, width=10)
column_width_entry.grid(row=2, column=1, padx=5, pady=5)
column_width_entry.insert(0, str(config["column_width"]))

tk.Label(frame, text="行高:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
row_height_entry = tk.Entry(frame, width=10)
row_height_entry.grid(row=3, column=1, padx=5, pady=5)
row_height_entry.insert(0, str(config["row_height"]))

update_button = tk.Button(root, text="更新配置", command=update_config)
update_button.pack(pady=10)

# 开始按钮
start_button = tk.Button(root, text="开始处理", command=start_processing, bg="green", fg="white")
start_button.pack(pady=20)

# 运行GUI
root.mainloop()

