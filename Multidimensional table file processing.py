from PIL import Image, JpegImagePlugin
JpegImagePlugin._getmp = lambda x: None  # 忽略MPO多图片对象

import pandas as pd
import os
import shutil
import zipfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import tkinter as tk
from tkinter import filedialog, messagebox

# 默认配置
config = {
    "image_width": 70,        # 图片宽度(px)
    "image_height": 30,       # 图片高度(px)
    "column_width": 15,       # 列宽
    "row_height": 40          # 行高
}

def convert_image_to_supported(image_path):
    """
    转换图片为Excel支持的格式(PNG)
    """
    try:
        # 打开图片
        with Image.open(image_path) as img:
            # 获取图片格式
            img_format = img.format

            # 如果不是PNG或JPEG，转换为PNG
            if img_format not in ['PNG', 'JPEG']:
                png_path = os.path.splitext(image_path)[0] + '.png'
                img.convert('RGB').save(png_path, 'PNG')
                return png_path
            return image_path
    except Exception as e:
        print(f"图片转换错误 {image_path}: {e}")
        return None

def find_file_paths(folder_path, file_names):
    """
    在文件夹及其子文件夹中查找多个文件路径
    """
    found_files = {}
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file in file_names:
                full_path = os.path.join(root, file)
                if is_valid_file(full_path):
                    found_files[file] = full_path
    return found_files

def is_valid_file(file_path):
    """
    检查文件是否为有效的图片或其他支持的文件类型
    """
    valid_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.zip']
    file_extension = os.path.splitext(file_path)[1].lower()
    return file_extension in valid_extensions

def adjust_cell_size(ws, col, row):
    """
    根据用户配置调整单元格尺寸
    """
    ws.column_dimensions[col].width = config["column_width"]
    ws.row_dimensions[row].height = config["row_height"]

def insert_image_to_cell(ws, row, col, image_path):
    """
    在指定单元格中插入图片
    """
    try:
        # 转换图片为支持的格式
        converted_image_path = convert_image_to_supported(image_path)

        if not converted_image_path:
            return

        img = OpenPyXLImage(converted_image_path)
        # 根据用户设置调整图片大小
        img.width, img.height = config["image_width"], config["image_height"]

        col_letter = get_column_letter(col)
        adjust_cell_size(ws, col_letter, row)  # 调整单元格大小

        ws.add_image(img, f"{col_letter}{row}")

        # 清空单元格原有文字
        ws.cell(row=row, column=col).value = ''
    except Exception as e:
        print(f"无法插入图片 {image_path}: {e}")

def create_attachments_folder(excel_path):
    """
    创建附件文件夹
    """
    attachments_folder = os.path.join(os.path.dirname(excel_path), "附件")
    if not os.path.exists(attachments_folder):
        os.makedirs(attachments_folder)
    return attachments_folder

def copy_files_to_attachments(attachments_folder, images, files, zip_counter):
    """
    将图片和文件打包到一个zip文件中
    """
    zip_name = f"附件{zip_counter}.zip"
    zip_path = os.path.join(attachments_folder, zip_name)
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        # 打包图片
        for image in images:
            zipf.write(image, arcname=os.path.basename(image))

        # 打包其他文件
        for file in files:
            zipf.write(file, arcname=os.path.basename(file))

    return zip_name

def insert_hyperlink(ws, row, col, display_text, target):
    """
    在指定单元格中插入超链接并清空原有文字
    """
    cell = ws.cell(row=row, column=col)
    cell.value = display_text
    cell.hyperlink = target
    cell.style = "Hyperlink"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(color="0000FF", underline="single")  # 蓝色下划线

    col_letter = get_column_letter(col)
    adjust_cell_size(ws, col_letter, row)

def replace_and_insert_files(excel_path, folder_path):
    """
    替换Excel中的文件名为路径，并插入对应的图片或超链接
    """
    output_path = os.path.join(os.path.dirname(excel_path), "Updated_" + os.path.basename(excel_path))
    print(f"保存的Excel文件路径: {output_path}")
    try:
        df = pd.read_excel(excel_path)
        wb = load_workbook(excel_path)
        ws = wb.active

        attachments_folder = create_attachments_folder(excel_path)
        zip_counter = 1

        for col_index, col in enumerate(df.columns, start=1):
            for row_index, cell_value in enumerate(df[col], start=2):
                if not isinstance(cell_value, str):
                    continue

                file_names = [name.strip() for name in cell_value.split(',')]
                if not file_names:
                    continue

                found_files = find_file_paths(folder_path, file_names)
                if not found_files:
                    continue

                images = []
                other_files = []

                for file_name in file_names:
                    if file_name in found_files:
                        file_path = found_files[file_name]
                        if is_image_file(file_path):
                            images.append(file_path)
                        else:
                            other_files.append(file_path)

                # 如果同时存在图片和其他文件，打包它们
                if images or other_files:
                    # 清空单元格原有文字
                    ws.cell(row=row_index, column=col_index).value = ''

                # 处理图片和其他文件
                if images or other_files:
                    zip_name = copy_files_to_attachments(attachments_folder, images, other_files, zip_counter)

                    # 插入超链接并清空单元格
                    relative_path = os.path.join("附件", zip_name)
                    insert_hyperlink(ws, row_index, col_index, f"附件{zip_counter}", relative_path)

                    zip_counter += 1

        wb.save(output_path)
        print(f"Excel文件已保存: {output_path}")
        messagebox.showinfo("完成", f"处理完成！文件已保存到：\n{output_path}")
    except Exception as e:
        messagebox.showerror("处理错误", f"处理时发生错误：\n{e}")

def is_image_file(file_path):
    """判断文件是否为图片"""
    valid_image_extensions = ['.jpg', '.jpeg', '.png', '.webp']
    file_extension = os.path.splitext(file_path)[1].lower()
    return file_extension in valid_image_extensions

def select_excel():
    """选择Excel文件"""
    global excel_path
    excel_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel文件", "*.xlsx")])
    excel_label.config(text=f"已选择: {excel_path}")

def select_folder():
    """选择图片及其他文件夹"""
    global folder_path
    folder_path = filedialog.askdirectory(title="选择包含图片及其他文件的文件夹")
    folder_label.config(text=f"已选择: {folder_path}")

def update_config_values():
    """更新配置"""
    try:
        config["image_width"] = int(image_width_entry.get())
        config["image_height"] = int(image_height_entry.get())
        config["column_width"] = int(column_width_entry.get())
        config["row_height"] = int(row_height_entry.get())
        messagebox.showinfo("设置完成", "配置已更新！")
    except ValueError:
        messagebox.showerror("输入错误", "请输入有效的数字！")

def start_processing():
    """开始处理"""
    if not excel_path or not folder_path:
        messagebox.showwarning("警告", "请先选择Excel文件和图片及文件夹")
        return
    replace_and_insert_files(excel_path, folder_path)

# 创建主窗口
root = tk.Tk()
root.title("品质整改表整理")
root.geometry("600x600")

# Excel选择
excel_label = tk.Label(root, text="请选择Excel文件")
excel_label.pack(pady=10)
excel_button = tk.Button(root, text="选择Excel文件", command=select_excel)
excel_button.pack()

# 文件夹选择
folder_label = tk.Label(root, text="请选择图片及文件所在文件夹")
folder_label.pack(pady=10)
folder_button = tk.Button(root, text="选择文件夹", command=select_folder)
folder_button.pack()

# 设置区域
settings_label = tk.Label(root, text="设置图片和表格尺寸", font=("Arial", 12))
settings_label.pack(pady=10)

frame = tk.Frame(root)
frame.pack(pady=5)

# 图片宽度
tk.Label(frame, text="图片宽度(px):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
image_width_entry = tk.Entry(frame, width=10)
image_width_entry.grid(row=0, column=1, padx=5, pady=5)
image_width_entry.insert(0, str(config["image_width"]))

# 图片高度
tk.Label(frame, text="图片高度(px):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
image_height_entry = tk.Entry(frame, width=10)
image_height_entry.grid(row=1, column=1, padx=5, pady=5)
image_height_entry.insert(0, str(config["image_height"]))

# 列宽
tk.Label(frame, text="列宽:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
column_width_entry = tk.Entry(frame, width=10)
column_width_entry.grid(row=4, column=1, padx=5, pady=5)
column_width_entry.insert(0, str(config["column_width"]))

# 行高
tk.Label(frame, text="行高:").grid(row=5, column=0, padx=5, pady=5, sticky="e")
row_height_entry = tk.Entry(frame, width=10)
row_height_entry.grid(row=5, column=1, padx=5, pady=5)
row_height_entry.insert(0, str(config["row_height"]))

update_button = tk.Button(root, text="更新配置", command=update_config_values)
update_button.pack(pady=10)

# 开始按钮
start_button = tk.Button(root, text="开始处理", command=start_processing, bg="green", fg="white", font=("Arial", 12))
start_button.pack(pady=20)

# 初始化全局变量
excel_path = ""
folder_path = ""

# 运行GUI
root.mainloop()
